import os.path
import random
import openpyxl
import functools


def mscore(avg, num, max_delta=7):
    scores, retries = [], 5
    while retries:
        bias = 0
        for i in range(1, num):
            bias = int(random.uniform(-3-bias, 3-bias))
            scores.append(avg-bias)
        scores.append(avg*num - sum(scores))
        delta = max([ abs(avg-v) for v in scores])
        if delta <= max_delta:
            return scores
        retries = retries -1
    return [avg]*num


def sscore(avg, weights, max_delta=7):
    scores, retries = [], 5
    while retries:
        bias = 0
        for w in weights[:-1]:
            cscore = int(avg*w/100)
            bias = int(random.uniform(-3-bias, 3-bias))
            scores.append(cscore-bias)
        scores.append(avg - sum(scores))
        delta = max([abs(avg*w/100-s) for s,w in zip(scores, weights)])
        if delta <= max_delta:
            return scores
        retries = retries -1
    scores = [int(avg * w / 100) for w in weights[:-1]]
    scores.append(avg - sum(scores))
    return scores


def find(sheet, value):
    for r in range(sheet.min_row, sheet.max_row):
        for c in range(sheet.min_column, sheet.max_column):
            if sheet.cell(r,c).value==value:
                return r,c


def read_xls(fn,rscore, arg="实验报告",num=2):
    wb = openpyxl.load_workbook(fn,keep_vba=True, data_only=True)
    sheets = wb.worksheets  # 获取当前所有的sheet
    # 获取第一张sheet
    sheet1 = sheets[0]
    sidx, cno = find(sheet1, "学号")
    sidx1, cname = find(sheet1, "姓名")
    sidx2, ctotal = find(sheet1, arg)
    assert(sidx == sidx1)
    rows = []
    for r in range(sidx+1, sheet1.max_row+1):
        sno = sheet1.cell(r,cno).value
        sname = sheet1.cell(r,cname).value
        stotal = sheet1.cell(r,ctotal).value
        if not sno:
            break
        else:
            rows.append([sno, sname]+rscore(int(stotal))+[stotal])
    return rows


def write_xls(rows, fn,  num=2):
    wb = openpyxl.Workbook()
    # 获取当前活跃的sheet，默认是第一个sheet
    ws = wb.active
    ws['A1'] = '学号'
    ws['B1'] = '姓名'
    for i in range(num):
        cc = chr(ord("C")+i)
        ws[f'{cc}1'].value = f'分数{i+1}'
    cc = chr(ord("C") + num)
    ws[f'{cc}1'] = '合计'
    for r in rows:
        r = [str(v) for v in r]
        ws.append(r)
    for r in range(ws.min_row, ws.max_row+1):
        for c in range(ws.min_column, ws.max_column+1):
            ws.cell(row=r, column=c).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    wb.save(fn)


def main(src,rscore, strg="9、实验报告成绩登记表.xlsx",arg="实验报告", num=2):
    basedir = os.path.dirname(os.path.abspath(src))
    trg = os.path.abspath(os.path.join(basedir, strg))
    rows = read_xls(fn=src,rscore=rscore,arg=arg,num=num)
    write_xls(rows, fn=trg,num=num)

def mmain():
    lst = [61, 65, 67, 72, 75, 78, 81, 85, 92, 95]
    for v in lst:
        print(v, mscore(v, 2))
        print(v, mscore(v, 5))

def smain():
    lst = [61, 65, 67, 72, 75, 78, 81, 85, 92, 95]
    weights = [25,15,60]
    for v in lst:
        print(v, sscore(v, weights))
        print(v, sscore(v, weights))


def rmain():
    # lst = ["C://Users//jinya//Desktop//2024_2成绩汇总//情感分析//21级智能1班//6、课程过程考核成绩登记表.xlsx",
    #        "C://Users//jinya//Desktop//2024_2成绩汇总//情感分析//21级智能2班//6、课程过程考核成绩登记表.xlsx",
    #        "C://Users//jinya//Desktop//2024_2成绩汇总//人机对话//21级智能1班//6、课程过程考核成绩登记表.xlsx",
    #        "C://Users//jinya//Desktop//2024_2成绩汇总//人机对话//21级智能2班//6、课程过程考核成绩登记表.xlsx",]
    # rscore = functools.partial(mscore, num=2)
    # for fn in lst:
    #     main(fn,rscore=rscore)
    # fn="C://Users//jinya//Desktop//2024_2成绩汇总//智能前沿//6、课程过程考核成绩登记表.xlsx"
    # rscore = functools.partial(mscore, num=5)
    # main(fn,rscore=rscore,num=5)

    lst = [
        # "C://Users//jinya//Desktop//2024_2成绩汇总//情感分析//21级智能1班//11、学生期末考核大作业登记表及成绩.xlsx",
        #    "C://Users//jinya//Desktop//2024_2成绩汇总//情感分析//21级智能2班//11、学生期末考核大作业登记表及成绩.xlsx",
        #    "C://Users//jinya//Desktop//2024_2成绩汇总//人机对话//21级智能1班//11、学生期末考核大作业登记表及成绩.xlsx",
           "C://Users//jinya//Desktop//2024_2成绩汇总//人机对话//21级智能2班//11、学生期末考核大作业登记表及成绩.xlsx",
           ]
    weights = [25, 25, 60]
    rscore = functools.partial(sscore, weights=weights)
    for fn in lst:
        main(fn,rscore=rscore, strg="11、期末考核大作业成绩.xlsx", arg="合计", num=len(weights))
    fn = "C://Users//jinya//Desktop//2024_2成绩汇总//智能前沿//11、学生期末考核大作业登记表及成绩.xlsx"
    main(fn, rscore=rscore,strg="11、期末考核大作业成绩.xlsx", arg="合计", num=len(weights))


if __name__ == '__main__':
    rmain()


